import logging
from fastapi import FastAPI, HTTPException, Request, Depends, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from ad_auth import authenticate_user
from config import SECRET_KEY, ALGORITHM, ACCESS_TOKEN_EXPIRE_MINUTES
import jwt
import datetime
import sqlite3
from datetime import datetime as dt
from datetime import timedelta
import os
import shutil
import json
import urllib.parse
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import csv
from io import StringIO, BytesIO
from dotenv import load_dotenv
import calendar
import asyncio

# ============ ЗАГРУЗКА ПЕРЕМЕННЫХ ОКРУЖЕНИЯ ============
load_dotenv()

# ============ НАСТРОЙКИ ПОЧТЫ (из .env) ============
SMTP_SERVER = os.getenv("SMTP_SERVER", "192.168.168.206")
SMTP_PORT = int(os.getenv("SMTP_PORT", 25))
SMTP_USER = os.getenv("SMTP_USER", "web-mail@gap-rt.ru")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "V9tSKNSdBS5dCfOskdwI")

logger = logging.getLogger("main")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

app = FastAPI(title="Corporate Portal API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()

# ============ НАСТРОЙКА ДЛЯ НОВОСТЕЙ ============
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

NEWS_DB_PATH = os.path.join(BASE_DIR, "news.db")
CHAT_DB_PATH = os.path.join(BASE_DIR, "chat.db")
SETTINGS_DB_PATH = os.path.join(BASE_DIR, "settings.db")

# ============ ФУНКЦИИ ДЛЯ ОТПРАВКИ ПИСЕМ ============

def get_event_type_name(event_type):
    types = {
        'meeting': 'Совещание',
        'vks': 'ВКС',
        'deadline': 'Задача',
        'replacement': 'Замена'
    }
    return types.get(event_type, event_type)

def send_calendar_event_email(to_email, to_name, event, is_new=True):
    if not to_email:
        logger.warning(f"Email не указан для {to_name}")
        return False
    
    try:
        event_date = event.get('event_date')
        if event_date and '-' in event_date:
            event_date_display = f"{event_date[8:10]}.{event_date[5:7]}.{event_date[0:4]}"
        else:
            event_date_display = event_date
        
        event_time = event.get('event_time', '10:00') if not event.get('is_all_day') else 'весь день'
        
        if is_new:
            subject = f"📅 Новое событие в календаре: {event.get('title')}"
        else:
            subject = f"✏️ Изменение события в календаре: {event.get('title')}"
        
        event_type_name = get_event_type_name(event.get('event_type', 'meeting'))
        
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                <div style="background: linear-gradient(135deg, #3b82f6, #4f46e5); padding: 20px; color: white; text-align: center;">
                    <h2 style="margin: 0; font-size: 24px;">📅 {event.get('title')}</h2>
                </div>
                <div style="padding: 20px;">
                    <p>Уважаемый(ая) <strong>{to_name}</strong>!</p>
                    <p>Вас добавили в событие календаря.</p>
                    <table style="width: 100%; border-collapse: collapse; margin: 16px 0;">
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📅 Дата:</strong></td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event_date_display} {f'в {event_time}' if event_time != 'весь день' else '(весь день)'}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>🏷️ Тип:</strong></td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event_type_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📍 Место:</strong></td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event.get('location') or 'Не указано'}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📝 Описание:</strong></td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event.get('description') or 'Нет'}</td>
                        </tr>
                    </table>
                    <hr>
                    <p style="text-align: center;">
                        <a href="https://srv-app16.gap-rt.ru" style="display: inline-block; background: #3b82f6; color: white; padding: 10px 20px; text-decoration: none; border-radius: 8px;">Перейти в календарь</a>
                    </p>
                    <p style="color: #666; font-size: 12px; text-align: center;">Это автоматическое уведомление. Пожалуйста, не отвечайте на него.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        if SMTP_USER and SMTP_PASSWORD:
            server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        logger.info(f"✅ Email о событии отправлен {to_name} на {to_email}")
        return True
    except Exception as e:
        logger.error(f"❌ Ошибка отправки email о событии: {e}")
        return False

def send_task_email(to_email, to_name, task_title, task_description, due_date, task_id):
    if not to_email:
        logger.warning(f"Email не указан для исполнителя {to_name}")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = to_email
        msg['Subject'] = f"📋 Новая IT-задача: {task_title}"
        
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
        <h2>Уважаемый(ая) {to_name}!</h2>
        <p>Вам назначена новая задача в корпоративном портале.</p>
        <hr>
        <h3>📌 Задача: {task_title}</h3>
        <p><strong>📝 Описание:</strong> {task_description or 'Не указано'}</p>
        <p><strong>⏰ Срок выполнения:</strong> {due_date or 'Не указан'}</p>
        <hr>
        <p>Ссылка на портал: <a href="https://srv-app16.gap-rt.ru">Портал gap-rt.ru</a></p>
        <p>ID задачи: #{task_id}</p>
        <hr>
        <p style="color: #666; font-size: 12px;">Это автоматическое уведомление. Пожалуйста, не отвечайте на него.</p>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        if SMTP_USER and SMTP_PASSWORD:
            server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        logger.info(f"✅ Email отправлен исполнителю {to_name} на {to_email}")
        return True
    except Exception as e:
        logger.error(f"❌ Ошибка отправки email исполнителю: {e}")
        return False

def send_event_notification_to_all_participants(event_id, conn, is_new=True):
    logger.info(f"========== ОТПРАВКА EMAIL ДЛЯ СОБЫТИЯ {event_id} ==========")
    c = conn.cursor()
    
    c.execute("SELECT * FROM calendar_events WHERE id = ?", (event_id,))
    event = dict(c.fetchone())
    logger.info(f"Событие: {event.get('title')}")
    logger.info(f"Дата: {event.get('event_date')}")
    logger.info(f"Время: {event.get('event_time')}")
    
    c.execute("SELECT participant_type, participant_id FROM calendar_event_participants WHERE event_id = ?", (event_id,))
    participants = c.fetchall()
    logger.info(f"Найдено участников: {len(participants)}")
    
    chat_conn = sqlite3.connect(CHAT_DB_PATH)
    chat_conn.row_factory = sqlite3.Row
    chat_c = chat_conn.cursor()
    
    for p_type, p_id in participants:
        logger.info(f"Обработка участника: тип={p_type}, id={p_id}")
        if p_type == 'user':
            chat_c.execute("SELECT name, email FROM users WHERE username = ?", (p_id,))
            user = chat_c.fetchone()
            if user and user['email']:
                send_calendar_event_email(user['email'], user['name'] or p_id, event, is_new)
            else:
                logger.warning(f"Не найден email для пользователя {p_id}")
        elif p_type == 'calendar_group':
            c.execute("SELECT username FROM calendar_group_members WHERE group_id = ?", (p_id,))
            members = c.fetchall()
            for member in members:
                chat_c.execute("SELECT name, email FROM users WHERE username = ?", (member[0],))
                user = chat_c.fetchone()
                if user and user['email']:
                    send_calendar_event_email(user['email'], user['name'] or member[0], event, is_new)
    
    chat_conn.close()
    logger.info(f"========== ОТПРАВКА EMAIL ДЛЯ СОБЫТИЯ {event_id} ЗАВЕРШЕНА ==========")

# ============ НАПОМИНАНИЯ И ТОСТЫ ============

async def create_toast_notification(user_id, event):
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS toast_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT NOT NULL,
        event_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        event_date TEXT,
        event_time TEXT,
        location TEXT,
        description TEXT,
        event_type TEXT,
        remind_before INTEGER,
        shown INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (event_id) REFERENCES calendar_events(id) ON DELETE CASCADE
    )''')
    conn.commit()
    
    c.execute("""
        INSERT INTO toast_notifications (user_id, event_id, title, event_date, event_time, location, description, event_type, remind_before)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (user_id, event['id'], event.get('title'), event.get('event_date'), event.get('event_time'), 
          event.get('location'), event.get('description'), event.get('event_type'), event.get('remind_before')))
    conn.commit()
    conn.close()

async def send_reminder_to_participants(event, participants):
    logger.info(f"📧 Отправка напоминания для события: {event.get('title')}")
    chat_conn = get_db()
    chat_c = chat_conn.cursor()
    
    users_data = {}
    for row in chat_c.execute("SELECT username, name, email FROM users WHERE email IS NOT NULL AND email != ''"):
        users_data[row['username']] = {'name': row['name'], 'email': row['email']}
    chat_conn.close()
    
    event_date = event.get('event_date')
    if event_date and '-' in event_date:
        event_date_display = f"{event_date[8:10]}.{event_date[5:7]}.{event_date[0:4]}"
    else:
        event_date_display = event_date
    
    event_time = event.get('event_time', '10:00') if not event.get('is_all_day') else 'весь день'
    remind_before = event.get('remind_before', 0)
    
    if remind_before >= 1440:
        remind_text = f"за {remind_before // 1440} дней"
    elif remind_before >= 60:
        remind_text = f"за {remind_before // 60} часов"
    else:
        remind_text = f"за {remind_before} минут"
    
    event_type_name = get_event_type_name(event.get('event_type', 'meeting'))
    
    for p in participants:
        if p['participant_type'] == 'user':
            username = p['participant_id']
            if username in users_data:
                user = users_data[username]
                
                subject = f"🔔 Напоминание о событии: {event.get('title')}"
                body = f"""<html>
<body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px;">
    <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
        <div style="background: linear-gradient(135deg, #f59e0b, #d97706); padding: 20px; color: white; text-align: center;">
            <h2 style="margin: 0; font-size: 24px;">🔔 {event.get('title')}</h2>
        </div>
        <div style="padding: 20px;">
            <p>Уважаемый(ая) <strong>{user['name'] or username}</strong>!</p>
            <p>Напоминаем вам о предстоящем событии <strong>{remind_text}</strong>.</p>
            <table style="width: 100%; border-collapse: collapse; margin: 16px 0;">
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📅 Дата:</strong></td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event_date_display} {f'в {event_time}' if event_time != 'весь день' else '(весь день)'}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>🏷️ Тип:</strong></td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event_type_name}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📍 Место:</strong></td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event.get('location') or 'Не указано'}</td>
                </tr>
                <tr>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;"><strong>📝 Описание:</strong></td>
                    <td style="padding: 10px 0; border-bottom: 1px solid #e2e8f0;">{event.get('description') or 'Нет'}</td>
                </tr>
            </table>
            <hr>
            <p style="text-align: center;">
                <a href="https://srv-app16.gap-rt.ru" style="display: inline-block; background: #f59e0b; color: white; padding: 10px 20px; text-decoration: none; border-radius: 8px;">Перейти в календарь</a>
            </p>
            <p style="color: #666; font-size: 12px; text-align: center;">Это автоматическое напоминание. Пожалуйста, не отвечайте на него.</p>
        </div>
    </div>
</body>
</html>"""
                
                try:
                    msg = MIMEMultipart()
                    msg['From'] = SMTP_USER
                    msg['To'] = user['email']
                    msg['Subject'] = subject
                    msg.attach(MIMEText(body, 'html'))
                    
                    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
                    if SMTP_USER and SMTP_PASSWORD:
                        server.login(SMTP_USER, SMTP_PASSWORD)
                    server.send_message(msg)
                    server.quit()
                    logger.info(f"✅ Email напоминания отправлен {username} на {user['email']}")
                except Exception as e:
                    logger.error(f"❌ Ошибка отправки email {username}: {e}")
                
                await create_toast_notification(username, event)
                logger.info(f"✅ Тост-уведомление создано для {username}")

async def check_and_send_reminders():
    while True:
        try:
            now_local = dt.now()
            current_time = now_local.strftime('%H:%M')
            today = now_local.strftime('%Y-%m-%d')
            
            # Логируем только если есть события (не каждую минуту)
            conn = get_settings_db()
            c = conn.cursor()
            
            # Получаем события на сегодня с напоминанием
            c.execute("""
                SELECT id, title, event_time, remind_before 
                FROM calendar_events 
                WHERE event_date = ? AND remind_before > 0
            """, (today,))
            
            events = c.fetchall()
            
            for event_id, title, event_time, remind_before in events:
                # Проверяем, не отправляли ли уже
                c.execute("SELECT id FROM reminder_sent WHERE event_id = ? AND remind_before = ?", (event_id, remind_before))
                if c.fetchone():
                    continue  # Уже отправлено, пропускаем
                
                # Вычисляем время напоминания
                hour = int(event_time.split(':')[0])
                minute = int(event_time.split(':')[1]) - remind_before
                if minute < 0:
                    hour -= 1
                    minute += 60
                reminder_time = f"{hour:02d}:{minute:02d}"
                
                # Если время пришло
                if reminder_time <= current_time:
                    logger.info(f"📢 Отправляем напоминание для: {title}")
                    
                    c.execute("SELECT participant_type, participant_id FROM calendar_event_participants WHERE event_id = ?", (event_id,))
                    participants = c.fetchall()
                    
                    event_dict = {
                        'id': event_id, 
                        'title': title, 
                        'event_time': event_time, 
                        'remind_before': remind_before,
                        'event_date': today,
                        'is_all_day': 0
                    }
                    
                    await send_reminder_to_participants(event_dict, participants)
                    
                    # Отмечаем что отправили
                    c.execute("INSERT INTO reminder_sent (event_id, remind_before) VALUES (?, ?)", (event_id, remind_before))
                    conn.commit()
                    logger.info(f"✅ Напоминание для {title} отправлено")
            
            conn.close()
        except Exception as e:
            logger.error(f"Ошибка: {e}")
        
        await asyncio.sleep(60)

# ============ ИНИЦИАЛИЗАЦИЯ БАЗ ДАННЫХ ============

def init_news_db():
    conn = sqlite3.connect(NEWS_DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS news (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        content TEXT NOT NULL,
        category TEXT DEFAULT 'announcement',
        author TEXT,
        pub_date TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS news_images (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        news_id INTEGER NOT NULL,
        image_url TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (news_id) REFERENCES news(id) ON DELETE CASCADE
    )''')
    conn.commit()
    conn.close()

def init_chat_db():
    conn = sqlite3.connect(CHAT_DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, name TEXT, department TEXT, email TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS chats (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, created_by TEXT NOT NULL, is_general INTEGER DEFAULT 0, created_at TEXT NOT NULL)")
    c.execute("CREATE TABLE IF NOT EXISTS chat_members (chat_id INTEGER, username TEXT, PRIMARY KEY (chat_id, username))")
    c.execute("CREATE TABLE IF NOT EXISTS messages (id INTEGER PRIMARY KEY AUTOINCREMENT, chat_id INTEGER, sender_id TEXT, sender_name TEXT, text TEXT, is_file INTEGER DEFAULT 0, file_name TEXT, file_data TEXT, file_type TEXT, edited INTEGER DEFAULT 0, reply_to INTEGER, timestamp TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS read_receipts (chat_id INTEGER, username TEXT, last_read_time TEXT, PRIMARY KEY (chat_id, username))")
    general = c.execute("SELECT id FROM chats WHERE is_general = 1").fetchone()
    if not general:
        c.execute("INSERT INTO chats (name, created_by, is_general, created_at) VALUES (?, ?, ?, ?)", ("Общий чат", "system", 1, dt.now().isoformat()))
    conn.commit()
    conn.close()

def init_settings_db():
    conn = sqlite3.connect(SETTINGS_DB_PATH)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS ad_groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_name TEXT UNIQUE NOT NULL,
        display_name TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS resource_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        icon TEXT DEFAULT '📁',
        is_global INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS category_targets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category_id INTEGER NOT NULL,
        target_type TEXT NOT NULL,
        target_id TEXT NOT NULL,
        FOREIGN KEY (category_id) REFERENCES resource_categories(id) ON DELETE CASCADE
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS network_resources (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        resource_name TEXT NOT NULL,
        resource_path TEXT NOT NULL,
        resource_type TEXT DEFAULT 'folder',
        category_id INTEGER,
        is_global INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (category_id) REFERENCES resource_categories(id) ON DELETE SET NULL
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS network_resource_targets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        resource_id INTEGER NOT NULL,
        target_type TEXT NOT NULL,
        target_id TEXT NOT NULL,
        FOREIGN KEY (resource_id) REFERENCES network_resources(id) ON DELETE CASCADE
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_name TEXT NOT NULL,
        service_url TEXT NOT NULL,
        service_icon TEXT DEFAULT '🔗',
        is_global INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS service_targets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER NOT NULL,
        target_type TEXT NOT NULL,
        target_id TEXT NOT NULL,
        FOREIGN KEY (service_id) REFERENCES services(id) ON DELETE CASCADE
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        role_name TEXT UNIQUE NOT NULL,
        display_name TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        permission_key TEXT UNIQUE NOT NULL,
        display_name TEXT,
        module TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS role_permissions (
        role_id INTEGER,
        permission_id INTEGER,
        FOREIGN KEY (role_id) REFERENCES roles(id),
        FOREIGN KEY (permission_id) REFERENCES permissions(id),
        PRIMARY KEY (role_id, permission_id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS user_roles (
        username TEXT NOT NULL,
        role_id INTEGER,
        assigned_by TEXT,
        assigned_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (role_id) REFERENCES roles(id),
        PRIMARY KEY (username, role_id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS vacation_replacements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_name TEXT NOT NULL,
        position TEXT,
        department TEXT,
        substitute_name TEXT NOT NULL,
        start_date TEXT,
        end_date TEXT,
        reason TEXT,
        status TEXT DEFAULT 'active',
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS calendar_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        event_date TEXT NOT NULL,
        event_time TEXT DEFAULT '10:00',
        end_date TEXT,
        end_time TEXT,
        event_type TEXT DEFAULT 'meeting',
        location TEXT,
        description TEXT,
        is_all_day INTEGER DEFAULT 0,
        repeat TEXT DEFAULT 'none',
        remind_before INTEGER DEFAULT 0,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS calendar_event_participants (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id INTEGER NOT NULL,
        participant_type TEXT NOT NULL,
        participant_id TEXT NOT NULL,
        FOREIGN KEY (event_id) REFERENCES calendar_events(id) ON DELETE CASCADE
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS calendar_groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS calendar_group_members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id INTEGER NOT NULL,
        username TEXT NOT NULL,
        FOREIGN KEY (group_id) REFERENCES calendar_groups(id) ON DELETE CASCADE
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        text TEXT NOT NULL,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS organization_tree (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tree_data TEXT NOT NULL,
        updated_by TEXT,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS task_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        color TEXT DEFAULT '#3b82f6',
        sort_order INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS equipment_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        category TEXT DEFAULT 'pc_component',
        unit TEXT DEFAULT 'шт',
        sort_order INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS it_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        category_id INTEGER,
        equipment_id INTEGER,
        assigned_to TEXT,
        components_status TEXT DEFAULT 'missing',
        executor TEXT,
        created_by TEXT,
        created_date TEXT,
        due_date TEXT,
        completed_date TEXT,
        is_archived INTEGER DEFAULT 0,
        assigned_email TEXT,
        monitoring_id INTEGER,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (category_id) REFERENCES task_categories(id) ON DELETE SET NULL,
        FOREIGN KEY (equipment_id) REFERENCES equipment_types(id) ON DELETE SET NULL
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS it_task_monitoring (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title_template TEXT NOT NULL,
        description_template TEXT,
        category_id INTEGER,
        equipment_id INTEGER,
        assigned_to TEXT,
        executor TEXT,
        interval_days INTEGER DEFAULT 30,
        is_active INTEGER DEFAULT 1,
        last_run TEXT,
        next_run TEXT,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (category_id) REFERENCES task_categories(id) ON DELETE SET NULL,
        FOREIGN KEY (equipment_id) REFERENCES equipment_types(id) ON DELETE SET NULL
    )''')
    
    c.execute("INSERT OR IGNORE INTO roles (role_name, display_name) VALUES ('it_engineer', 'IT-инженер')")
    
    it_role = c.execute("SELECT id FROM roles WHERE role_name = 'it_engineer'").fetchone()
    if it_role:
        it_permissions = [
            ('tasks.view', 'Просмотр задач', 'it_tasks'),
            ('tasks.create', 'Создание задач', 'it_tasks'),
            ('tasks.edit', 'Редактирование задач', 'it_tasks'),
            ('tasks.delete', 'Удаление задач', 'it_tasks'),
            ('tasks.archive', 'Архивирование задач', 'it_tasks'),
            ('tasks.restore', 'Восстановление из архива', 'it_tasks'),
            ('equipment.view', 'Просмотр комплектующих', 'it_tasks'),
            ('equipment.manage', 'Управление комплектующими', 'it_tasks'),
            ('reports.view', 'Просмотр отчетов', 'reports'),
        ]
        for perm_key, display_name, module in it_permissions:
            c.execute("INSERT OR IGNORE INTO permissions (permission_key, display_name, module) VALUES (?, ?, ?)", 
                     (perm_key, display_name, module))
            perm = c.execute("SELECT id FROM permissions WHERE permission_key = ?", (perm_key,)).fetchone()
            if perm:
                c.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", 
                         (it_role[0], perm[0]))

    c.execute("SELECT COUNT(*) FROM task_categories")
    if c.fetchone()[0] == 0:
        default_categories = [
            ('🖥️ Сборка ПК', '#10b981'),
            ('🔧 Ремонт', '#f59e0b'),
            ('💿 Настройка ПО', '#3b82f6'),
            ('🌐 Сеть', '#8b5cf6'),
            ('🖨️ Принтеры/МФУ', '#ef4444'),
            ('🔒 Доступы/Права', '#ec4898'),
        ]
        for name, color in default_categories:
            c.execute("INSERT INTO task_categories (name, color) VALUES (?, ?)", (name, color))
    
    c.execute("SELECT COUNT(*) FROM equipment_types")
    if c.fetchone()[0] == 0:
        default_equipment = [
            ('Материнская плата', 'pc_component', 'шт'),
            ('Процессор', 'pc_component', 'шт'),
            ('Оперативная память (ОЗУ)', 'pc_component', 'ГБ'),
            ('SSD-накопитель', 'pc_component', 'шт'),
            ('HDD-накопитель', 'pc_component', 'шт'),
            ('Блок питания', 'pc_component', 'шт'),
            ('Видеокарта', 'pc_component', 'шт'),
            ('Корпус', 'pc_component', 'шт'),
            ('Картридж', 'printer', 'шт'),
            ('Барабан', 'printer', 'шт'),
            ('Роутер', 'network', 'шт'),
            ('Коммутатор', 'network', 'шт'),
        ]
        for name, category, unit in default_equipment:
            c.execute("INSERT INTO equipment_types (name, category, unit) VALUES (?, ?, ?)", (name, category, unit))

    c.execute("INSERT OR IGNORE INTO roles (role_name, display_name) VALUES ('admin', 'Администратор')")
    c.execute("INSERT OR IGNORE INTO roles (role_name, display_name) VALUES ('department_head', 'Начальник отдела')")
    c.execute("INSERT OR IGNORE INTO roles (role_name, display_name) VALUES ('moderator', 'Модератор')")
    c.execute("INSERT OR IGNORE INTO roles (role_name, display_name) VALUES ('user', 'Пользователь')")

    permissions = [
        ('news.create', 'Создание новостей', 'news'),
        ('news.edit', 'Редактирование новостей', 'news'),
        ('news.delete', 'Удаление новостей', 'news'),
        ('replacements.create', 'Создание замен на отпуск', 'hr'),
        ('replacements.edit', 'Редактирование замен на отпуск', 'hr'),
        ('replacements.delete', 'Удаление замен на отпуск', 'hr'),
        ('calendar.create', 'Создание событий календаря', 'calendar'),
        ('calendar.edit', 'Редактирование событий календаря', 'calendar'),
        ('calendar.delete', 'Удаление событий календаря', 'calendar'),
        ('services.manage', 'Управление сервисами', 'services'),
        ('network.manage', 'Управление сетевыми ресурсами', 'network'),
        ('categories.manage', 'Управление категориями', 'network'),
        ('groups.manage', 'Управление AD группами', 'groups'),
        ('users.manage', 'Управление пользователями', 'admin'),
        ('notifications.manage', 'Управление оповещениями', 'admin'),
        ('organization.manage', 'Управление структурой организации', 'organization'),
    ]
    for perm in permissions:
        c.execute("INSERT OR IGNORE INTO permissions (permission_key, display_name, module) VALUES (?, ?, ?)", perm)

    admin_role = c.execute("SELECT id FROM roles WHERE role_name = 'admin'").fetchone()
    if admin_role:
        all_perms = c.execute("SELECT id FROM permissions").fetchall()
        for perm in all_perms:
            c.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (admin_role[0], perm[0]))

    dept_role = c.execute("SELECT id FROM roles WHERE role_name = 'department_head'").fetchone()
    if dept_role:
        dept_perms = ['replacements.create', 'replacements.edit', 'replacements.delete', 'calendar.create', 'calendar.edit', 'calendar.delete']
        for perm_key in dept_perms:
            perm = c.execute("SELECT id FROM permissions WHERE permission_key = ?", (perm_key,)).fetchone()
            if perm:
                c.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (dept_role[0], perm[0]))

    mod_role = c.execute("SELECT id FROM roles WHERE role_name = 'moderator'").fetchone()
    if mod_role:
        mod_perms = ['news.create', 'news.edit', 'news.delete']
        for perm_key in mod_perms:
            perm = c.execute("SELECT id FROM permissions WHERE permission_key = ?", (perm_key,)).fetchone()
            if perm:
                c.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (mod_role[0], perm[0]))

    c.execute("INSERT OR IGNORE INTO ad_groups (group_name, display_name) VALUES ('!citovmt', 'IT отдел')")
    
    default_tree = [
        {
            "id": 1,
            "name": "IT отдел",
            "type": "department",
            "employee": "",
            "children": [
                {"id": 2, "name": "Начальник отдела", "type": "position", "employee": "Сенюшин Александр Вл.", "children": []},
                {"id": 3, "name": "Системный администратор", "type": "position", "employee": "Титов Евгений Алексеевич", "children": []}
            ]
        }
    ]
    
    existing = c.execute("SELECT id FROM organization_tree LIMIT 1").fetchone()
    if not existing:
        c.execute("INSERT INTO organization_tree (tree_data, updated_by, updated_at) VALUES (?, ?, ?)",
                  (json.dumps(default_tree), "system", dt.now().isoformat()))
        logger.info("Organization tree initialized with default data")

    conn.commit()
    conn.close()
    logger.info("Settings database initialized")

def migrate_db():
    conn = sqlite3.connect(SETTINGS_DB_PATH)
    c = conn.cursor()

    try:
        c.execute('''CREATE TABLE IF NOT EXISTS service_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            service_id INTEGER NOT NULL,
            target_type TEXT NOT NULL,
            target_id TEXT NOT NULL,
            FOREIGN KEY (service_id) REFERENCES services(id) ON DELETE CASCADE
        )''')
    except: pass

    try:
        c.execute('''CREATE TABLE IF NOT EXISTS network_resource_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resource_id INTEGER NOT NULL,
            target_type TEXT NOT NULL,
            target_id TEXT NOT NULL,
            FOREIGN KEY (resource_id) REFERENCES network_resources(id) ON DELETE CASCADE
        )''')
    except: pass

    try:
        c.execute('''CREATE TABLE IF NOT EXISTS category_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER NOT NULL,
            target_type TEXT NOT NULL,
            target_id TEXT NOT NULL,
            FOREIGN KEY (category_id) REFERENCES resource_categories(id) ON DELETE CASCADE
        )''')
    except: pass

    try:
        c.execute("ALTER TABLE network_resources ADD COLUMN created_by TEXT")
    except: pass

    try:
        c.execute("ALTER TABLE it_tasks ADD COLUMN equipment_id INTEGER")
        c.execute("ALTER TABLE it_tasks ADD COLUMN assigned_email TEXT")
        c.execute("ALTER TABLE it_tasks ADD COLUMN monitoring_id INTEGER")
    except: pass

    try:
        c.execute("ALTER TABLE calendar_events ADD COLUMN end_date TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE calendar_events ADD COLUMN end_time TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE calendar_events ADD COLUMN repeat TEXT DEFAULT 'none'")
    except: pass
    try:
        c.execute("ALTER TABLE calendar_events ADD COLUMN remind_before INTEGER DEFAULT 0")
    except: pass

    try:
        c.execute('''CREATE TABLE IF NOT EXISTS calendar_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS calendar_group_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            FOREIGN KEY (group_id) REFERENCES calendar_groups(id) ON DELETE CASCADE
        )''')
    except: pass

    try:
        c.execute('''CREATE TABLE IF NOT EXISTS it_task_monitoring (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title_template TEXT NOT NULL,
            description_template TEXT,
            category_id INTEGER,
            equipment_id INTEGER,
            assigned_to TEXT,
            executor TEXT,
            interval_days INTEGER DEFAULT 30,
            is_active INTEGER DEFAULT 1,
            last_run TEXT,
            next_run TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES task_categories(id) ON DELETE SET NULL,
            FOREIGN KEY (equipment_id) REFERENCES equipment_types(id) ON DELETE SET NULL
        )''')
    except: pass

    try:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
    except:
        pass

    conn.commit()
    conn.close()
    logger.info("Database migration completed")

def add_email_column():
    conn = sqlite3.connect(CHAT_DB_PATH)
    c = conn.cursor()
    try:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
        logger.info("Колонка email добавлена в таблицу users")
    except:
        pass
    conn.commit()
    conn.close()

init_news_db()
init_chat_db()
init_settings_db()
migrate_db()
add_email_column()

# ============ МОДЕЛИ ============
class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    token: str
    token_type: str
    user: dict

# ============ АУТЕНТИФИКАЦИЯ ============
def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except:
        raise HTTPException(status_code=401, detail="Invalid token")

def get_db():
    conn = sqlite3.connect(CHAT_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_settings_db():
    conn = sqlite3.connect(SETTINGS_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def is_admin_by_group(user_groups: list) -> bool:
    return "!citovmt" in user_groups

async def get_user_role(username: str) -> str:
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("""
        SELECT r.role_name FROM user_roles ur
        JOIN roles r ON ur.role_id = r.id
        WHERE ur.username = ?
    """, (username,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else "user"

async def has_permission(username: str, permission_key: str, user_groups: list = None) -> bool:
    if user_groups and is_admin_by_group(user_groups):
        return True
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("""
        SELECT r.role_name FROM user_roles ur
        JOIN roles r ON ur.role_id = r.id
        WHERE ur.username = ?
    """, (username,))
    role_row = c.fetchone()
    role = role_row[0] if role_row else "user"
    if role == 'admin' or role == 'it_engineer':
        conn.close()
        return True
    c.execute("""
        SELECT 1 FROM role_permissions rp
        JOIN roles r ON rp.role_id = r.id
        JOIN permissions p ON rp.permission_id = p.id
        WHERE r.role_name = ? AND p.permission_key = ?
    """, (role, permission_key))
    result = c.fetchone()
    conn.close()
    return result is not None

def get_user_email_by_name(username: str) -> str:
    from server.ad_users import get_all_ad_users
    try:
        all_users = get_all_ad_users()
        for user in all_users:
            if user.get('name') == username or user.get('username') == username:
                return user.get('email', '')
    except:
        pass
    return ""

# ============ ФУНКЦИЯ ДЛЯ ДИНАМИЧЕСКИХ ПОВТОРЯЮЩИХСЯ СОБЫТИЙ ============
def get_recurring_dates(event_date, repeat, start_date, end_date):
    if repeat == "none":
        return [event_date]
    
    try:
        start = dt.strptime(start_date, "%Y-%m-%d")
        end = dt.strptime(end_date, "%Y-%m-%d")
        current = dt.strptime(event_date, "%Y-%m-%d")
    except:
        return [event_date]
    
    dates = []
    if repeat == "daily":
        delta = timedelta(days=1)
        while current <= end:
            if current >= start:
                dates.append(current.strftime("%Y-%m-%d"))
            current += delta
            
    elif repeat == "weekly":
        delta = timedelta(weeks=1)
        while current <= end:
            if current >= start:
                dates.append(current.strftime("%Y-%m-%d"))
            current += delta
            
    elif repeat == "monthly":
        while current <= end:
            if current >= start:
                dates.append(current.strftime("%Y-%m-%d"))
            year = current.year + ((current.month) // 12)
            month = (current.month % 12) + 1
            if month == 1:
                year += 1
            day = min(current.day, calendar.monthrange(year, month)[1])
            current = current.replace(year=year, month=month, day=day)
    
    return dates if dates else [event_date]

# ============ ЛОГИН ============
@app.post("/api/auth/login", response_model=LoginResponse)
async def login(request: Request):
    body = await request.json()
    login_data = LoginRequest(**body)

    user_data = authenticate_user(login_data.username, login_data.password)
    if not user_data or user_data.get('status') != 'success':
        raise HTTPException(status_code=401, detail="Invalid credentials")

    user_info = user_data['user_info']
    username = user_info["username"]
    display_name = user_info.get("display_name", username)

    groups = user_data.get('groups', [])

    with get_db() as conn:
        try:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT")
        except:
            pass
        user_email = user_info.get("email", "")
        conn.execute("INSERT OR IGNORE INTO users (username, name, department, email) VALUES (?, ?, ?, ?)",
                    (username, display_name, "", user_email))
        conn.execute("UPDATE users SET email = ?, name = ? WHERE username = ?", (user_email, display_name, username))
        general = conn.execute("SELECT id FROM chats WHERE is_general = 1").fetchone()
        if general:
            conn.execute("INSERT OR IGNORE INTO chat_members (chat_id, username) VALUES (?, ?)",
                        (general["id"], username))
        conn.commit()

    conn = get_settings_db()
    c = conn.cursor()
    c.execute("SELECT role_id FROM user_roles WHERE username = ?", (username,))
    existing_role = c.fetchone()
    if not existing_role:
        c.execute("SELECT id FROM roles WHERE role_name = 'user'")
        user_role = c.fetchone()
        if user_role:
            c.execute("INSERT INTO user_roles (username, role_id, assigned_by) VALUES (?, ?, ?)",
                     (username, user_role[0], "system"))
            logger.info(f"Автоматически назначена роль 'Пользователь' для {username}")
    
    role = await get_user_role(username)
    conn.close()

    token = create_access_token(data={
        "sub": username,
        "name": display_name,
        "email": user_info.get("email", ""),
        "display_name": display_name,
        "groups": groups,
        "group": user_data.get('group'),
        "role": role
    })

    return {
        "token": token, 
        "token_type": "bearer", 
        "user": {
            "username": username, 
            "name": display_name,
            "email": user_info.get("email", ""), 
            "display_name": display_name, 
            "groups": groups, 
            "group": user_data.get('group'),
            "role": role
        }
    }

@app.get("/api/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    role = current_user.get("role") or await get_user_role(username)
    return {
        "username": username, 
        "email": current_user.get("email"), 
        "display_name": current_user.get("display_name") or current_user.get("name", username), 
        "groups": current_user.get("groups", []), 
        "group": current_user.get("group"),
        "role": role
    }

@app.get("/api/auth/verify")
async def verify_token(current_user: dict = Depends(get_current_user)):
    return {"valid": True, "user": {"username": current_user.get("sub"), "display_name": current_user.get("display_name"), "email": current_user.get("email"), "groups": current_user.get("groups", [])}}

@app.post("/api")
async def api_root(request: Request):
    return {"status": "ok", "message": "API is running"}

@app.post("/api/iframe-auth")
async def iframe_auth(request: Request):
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.replace("Bearer ", "")
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            return {"loginToken": token, "user": {"username": payload.get("sub"), "name": payload.get("display_name"), "email": payload.get("email")}}
        except:
            pass
    return {"status": "ok", "message": "Iframe auth endpoint"}

# ============ НОВОСТИ ============
@app.get("/api/news")
async def get_news(limit: int = 20):
    conn = sqlite3.connect(NEWS_DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM news ORDER BY id DESC LIMIT ?", (limit,))
    news_list = [dict(row) for row in c.fetchall()]
    for news in news_list:
        c.execute("SELECT image_url FROM news_images WHERE news_id = ? ORDER BY sort_order", (news['id'],))
        news['images'] = [row['image_url'] for row in c.fetchall()]
    conn.close()
    return {"news": news_list}

@app.post("/api/news")
async def create_news(
    title: str = Form(...),
    content: str = Form(...),
    category: str = Form("announcement"),
    images: list[UploadFile] = File(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "news.create", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = sqlite3.connect(NEWS_DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO news (title, content, category, author, pub_date) VALUES (?, ?, ?, ?, ?)",
              (title, content, category, current_user.get("display_name", username), dt.now().isoformat()))
    news_id = c.lastrowid
    if images:
        for idx, image in enumerate(images):
            if image and image.filename:
                ext = image.filename.split('.')[-1] if '.' in image.filename else 'jpg'
                filename = f"{dt.now().timestamp()}_{idx}.{ext}"
                filepath = os.path.join(UPLOAD_DIR, filename)
                with open(filepath, "wb") as f:
                    shutil.copyfileobj(image.file, f)
                c.execute("INSERT INTO news_images (news_id, image_url, sort_order) VALUES (?, ?, ?)", (news_id, f"/uploads/{filename}", idx))
    conn.commit()
    conn.close()
    return {"id": news_id, "message": "Новость добавлена"}

@app.put("/api/news/{news_id}")
async def update_news(
    news_id: int,
    title: str = Form(...),
    content: str = Form(...),
    category: str = Form("announcement"),
    images: list[UploadFile] = File(None),
    delete_images: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "news.edit", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = sqlite3.connect(NEWS_DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE news SET title=?, content=?, category=? WHERE id=?", (title, content, category, news_id))
    if delete_images:
        try:
            for img_url in json.loads(delete_images):
                old_path = os.path.join(UPLOAD_DIR, os.path.basename(img_url))
                if os.path.exists(old_path):
                    os.remove(old_path)
                c.execute("DELETE FROM news_images WHERE news_id = ? AND image_url = ?", (news_id, img_url))
        except: pass
    if images:
        c.execute("SELECT COALESCE(MAX(sort_order), -1) FROM news_images WHERE news_id = ?", (news_id,))
        max_order = c.fetchone()[0]
        for idx, image in enumerate(images):
            if image and image.filename:
                ext = image.filename.split('.')[-1] if '.' in image.filename else 'jpg'
                filename = f"{dt.now().timestamp()}_{max_order + idx + 1}.{ext}"
                filepath = os.path.join(UPLOAD_DIR, filename)
                with open(filepath, "wb") as f:
                    shutil.copyfileobj(image.file, f)
                c.execute("INSERT INTO news_images (news_id, image_url, sort_order) VALUES (?, ?, ?)", (news_id, f"/uploads/{filename}", max_order + idx + 1))
    conn.commit()
    conn.close()
    return {"message": "Новость обновлена"}

@app.delete("/api/news/{news_id}")
async def delete_news(news_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "news.delete", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = sqlite3.connect(NEWS_DB_PATH)
    c = conn.cursor()
    for img in c.execute("SELECT image_url FROM news_images WHERE news_id = ?", (news_id,)).fetchall():
        filepath = os.path.join(UPLOAD_DIR, os.path.basename(img[0]))
        if os.path.exists(filepath):
            os.remove(filepath)
    c.execute("DELETE FROM news_images WHERE news_id = ?", (news_id,))
    c.execute("DELETE FROM news WHERE id = ?", (news_id,))
    conn.commit()
    conn.close()
    return {"deleted": news_id}

# ============ ОПОВЕЩЕНИЯ ============
@app.get("/api/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("SELECT * FROM notifications ORDER BY created_at DESC")
    notifications = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"notifications": notifications}

@app.post("/api/admin/notifications")
async def add_notification(text: str = Form(...), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not is_admin_by_group(user_groups) and await get_user_role(username) != 'admin':
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO notifications (text, created_by, created_at) VALUES (?, ?, ?)", (text, username, dt.now().isoformat()))
    conn.commit()
    conn.close()
    return {"id": c.lastrowid, "message": "Оповещение добавлено"}

@app.delete("/api/admin/notifications/{notification_id}")
async def delete_notification(notification_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not is_admin_by_group(user_groups) and await get_user_role(username) != 'admin':
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM notifications WHERE id = ?", (notification_id,))
    conn.commit()
    conn.close()
    return {"message": "Оповещение удалено"}

# ============ ЗАМЕНЫ НА ОТПУСК ============
@app.get("/api/vacation-replacements")
async def get_vacation_replacements(current_user: dict = Depends(get_current_user)):
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("SELECT * FROM vacation_replacements ORDER BY start_date DESC")
    replacements = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"replacements": replacements}

@app.post("/api/vacation-replacements")
async def add_vacation_replacement(
    employee_name: str = Form(...),
    position: str = Form(...),
    department: str = Form(...),
    substitute_name: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    reason: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "replacements.create", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO vacation_replacements (employee_name, position, department, substitute_name, start_date, end_date, reason, status, created_by, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
              (employee_name, position, department, substitute_name, start_date, end_date, reason, 'active', username, dt.now().isoformat()))
    conn.commit()
    replacement_id = c.lastrowid
    conn.close()
    return {"id": replacement_id, "message": "Замена добавлена"}

@app.delete("/api/vacation-replacements/{replacement_id}")
async def delete_vacation_replacement(replacement_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "replacements.delete", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM vacation_replacements WHERE id = ?", (replacement_id,))
    conn.commit()
    conn.close()
    return {"message": "Замена удалена"}

# ============ КАЛЕНДАРЬ ============
def is_event_visible_for_user(event_id: int, username: str, user_groups: list, conn) -> bool:
    c = conn.cursor()
    c.execute("SELECT created_by FROM calendar_events WHERE id = ?", (event_id,))
    event = c.fetchone()
    if not event:
        return False
    if event[0] == username:
        return True
    c.execute("SELECT participant_type, participant_id FROM calendar_event_participants WHERE event_id = ?", (event_id,))
    for p_type, p_id in c.fetchall():
        if p_type == 'user' and p_id == username:
            return True
        elif p_type == 'group':
            group = c.execute("SELECT group_name FROM ad_groups WHERE id = ?", (p_id,)).fetchone()
            if group and group[0] in user_groups:
                return True
    return False

@app.get("/api/calendar/events")
async def get_calendar_events(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    
    if not start_date:
        start_date = (dt.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = (dt.now() + timedelta(days=90)).strftime("%Y-%m-%d")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    chat_conn = get_db()
    chat_c = chat_conn.cursor()
    users_map = {}
    for row in chat_c.execute("SELECT username, name FROM users"):
        users_map[row[0]] = row[1] if row[1] else row[0]
    chat_conn.close()
    
    c.execute("""
        SELECT id, title, event_date, event_time, end_date, end_time, event_type,
               location, description, is_all_day, repeat, remind_before, created_by, created_at
        FROM calendar_events
        ORDER BY event_date, event_time
    """)
    
    all_events = [dict(row) for row in c.fetchall()]
    
    result = []
    for event in all_events:
        if not is_event_visible_for_user(event['id'], username, user_groups, conn):
            continue
        
        participants_raw = c.execute("""
            SELECT participant_type, participant_id 
            FROM calendar_event_participants 
            WHERE event_id = ?
        """, (event['id'],)).fetchall()
        
        participants_with_names = []
        for p_type, p_id in participants_raw:
            if p_type == 'user':
                user_name = users_map.get(p_id, p_id)
                participants_with_names.append({'type': 'user', 'id': p_id, 'name': user_name})
            elif p_type == 'group':
                group = c.execute("SELECT display_name FROM ad_groups WHERE id = ?", (p_id,)).fetchone()
                group_name = group[0] if group else p_id
                participants_with_names.append({'type': 'group', 'id': p_id, 'name': f"👥 {group_name}"})
            elif p_type == 'calendar_group':
                group = c.execute("SELECT name FROM calendar_groups WHERE id = ?", (p_id,)).fetchone()
                group_name = group[0] if group else p_id
                participants_with_names.append({'type': 'calendar_group', 'id': p_id, 'name': f"📅 {group_name}"})
        
        event["participants"] = participants_with_names
        result.append(event)
    
    conn.close()
    return {"events": result}

@app.post("/api/calendar/events")
async def add_calendar_event(
    title: str = Form(...),
    event_date: str = Form(...),
    event_time: str = Form("10:00"),
    end_date: str = Form(None),
    end_time: str = Form(None),
    event_type: str = Form("meeting"),
    location: str = Form(None),
    description: str = Form(None),
    is_all_day: int = Form(0),
    repeat: str = Form("none"),
    remind_before: int = Form(0),
    participants: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "calendar.create", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    if is_all_day == 1:
        event_time = "00:00"
        end_time = "00:00"
    
    c.execute("""INSERT INTO calendar_events 
                (title, event_date, event_time, end_date, end_time, event_type, location, description, is_all_day, repeat, remind_before, created_by, created_at) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
              (title, event_date, event_time, end_date or event_date, end_time, event_type, location, description, is_all_day, repeat, remind_before, username, dt.now().isoformat()))
    event_id = c.lastrowid
    
    if participants:
        try:
            for p in json.loads(participants):
                c.execute("INSERT INTO calendar_event_participants (event_id, participant_type, participant_id) VALUES (?, ?, ?)", 
                         (event_id, p.get('type'), p.get('id')))
        except: pass
    
    conn.commit()
    
    send_event_notification_to_all_participants(event_id, conn, is_new=True)
    
    conn.close()
    
    return {"id": event_id, "message": "Событие добавлено"}

@app.put("/api/calendar/events/{event_id}")
async def update_calendar_event(
    event_id: int,
    title: str = Form(...),
    event_date: str = Form(...),
    event_time: str = Form("10:00"),
    end_date: str = Form(None),
    end_time: str = Form(None),
    event_type: str = Form("meeting"),
    location: str = Form(None),
    description: str = Form(None),
    is_all_day: int = Form(0),
    repeat: str = Form("none"),
    remind_before: int = Form(0),
    participants: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "calendar.edit", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    event = c.execute("SELECT created_by, remind_before FROM calendar_events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        raise HTTPException(status_code=404, detail="Событие не найдено")
    if event[0] != username and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Вы можете редактировать только свои события")
    
    if is_all_day == 1:
        event_time = "00:00"
        end_time = "00:00"
    
    c.execute("""UPDATE calendar_events 
                SET title=?, event_date=?, event_time=?, end_date=?, end_time=?, event_type=?, location=?, description=?, is_all_day=?, repeat=?, remind_before=?
                WHERE id=?""",
              (title, event_date, event_time, end_date or event_date, end_time, event_type, location, description, is_all_day, repeat, remind_before, event_id))
    
    c.execute("DELETE FROM calendar_event_participants WHERE event_id = ?", (event_id,))
    if participants:
        try:
            for p in json.loads(participants):
                c.execute("INSERT INTO calendar_event_participants (event_id, participant_type, participant_id) VALUES (?, ?, ?)", 
                         (event_id, p.get('type'), p.get('id')))
        except: pass
    
    conn.commit()
    
    send_event_notification_to_all_participants(event_id, conn, is_new=False)
    
    conn.close()
    
    return {"message": "Событие обновлено"}

@app.delete("/api/calendar/events/{event_id}")
async def delete_calendar_event(event_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "calendar.delete", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    event = c.execute("SELECT created_by FROM calendar_events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        raise HTTPException(status_code=404, detail="Событие не найдено")
    if event[0] != username and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Вы можете удалять только свои события")
    
    c.execute("DELETE FROM calendar_event_participants WHERE event_id = ?", (event_id,))
    c.execute("DELETE FROM calendar_events WHERE id = ?", (event_id,))
    conn.commit()
    conn.close()
    return {"message": "Событие удалено"}

# ============ ГРУППЫ ДЛЯ КАЛЕНДАРЯ ============
@app.get("/api/calendar/groups")
async def get_calendar_groups(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role != 'admin' and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Доступ только для администраторов")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    groups = [dict(row) for row in c.execute("""
        SELECT * FROM calendar_groups ORDER BY created_at DESC
    """).fetchall()]
    
    for group in groups:
        members = c.execute("""
            SELECT username FROM calendar_group_members WHERE group_id = ?
        """, (group['id'],)).fetchall()
        group['members'] = [m[0] for m in members]
        group['members_count'] = len(members)
    
    conn.close()
    return {"groups": groups}

@app.post("/api/calendar/groups")
async def create_calendar_group(
    name: str = Form(...),
    description: str = Form(None),
    members: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role != 'admin' and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Доступ только для администраторов")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("""
        INSERT INTO calendar_groups (name, description, created_by, created_at)
        VALUES (?, ?, ?, ?)
    """, (name, description, username, dt.now().isoformat()))
    
    group_id = c.lastrowid
    
    if members:
        try:
            member_list = json.loads(members)
            for member in member_list:
                c.execute("""
                    INSERT INTO calendar_group_members (group_id, username)
                    VALUES (?, ?)
                """, (group_id, member))
        except: pass
    
    conn.commit()
    conn.close()
    
    return {"id": group_id, "message": "Группа создана"}

@app.put("/api/calendar/groups/{group_id}")
async def update_calendar_group(
    group_id: int,
    name: str = Form(...),
    description: str = Form(None),
    members: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role != 'admin' and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Доступ только для администраторов")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("""
        UPDATE calendar_groups SET name=?, description=?
        WHERE id=?
    """, (name, description, group_id))
    
    c.execute("DELETE FROM calendar_group_members WHERE group_id = ?", (group_id,))
    
    if members:
        try:
            member_list = json.loads(members)
            for member in member_list:
                c.execute("""
                    INSERT INTO calendar_group_members (group_id, username)
                    VALUES (?, ?)
                """, (group_id, member))
        except: pass
    
    conn.commit()
    conn.close()
    
    return {"message": "Группа обновлена"}

@app.delete("/api/calendar/groups/{group_id}")
async def delete_calendar_group(group_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role != 'admin' and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Доступ только для администраторов")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("DELETE FROM calendar_group_members WHERE group_id = ?", (group_id,))
    c.execute("DELETE FROM calendar_groups WHERE id = ?", (group_id,))
    
    conn.commit()
    conn.close()
    
    return {"message": "Группа удалена"}

@app.get("/api/calendar/available-users")
async def get_available_users(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    c = conn.cursor()
    users = [dict(row) for row in c.execute("""
        SELECT username, name as display_name FROM users ORDER BY name
    """).fetchall()]
    conn.close()
    return {"users": users}

# ============ СТРУКТУРА ОРГАНИЗАЦИИ ============
@app.get("/api/admin/organization-tree")
async def get_organization_tree(current_user: dict = Depends(get_current_user)):
    conn = get_settings_db()
    c = conn.cursor()
    row = c.execute("SELECT tree_data, updated_by, updated_at FROM organization_tree ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    if row:
        try:
            return {"tree": json.loads(row[0]), "updated_by": row[1], "updated_at": row[2]}
        except:
            return {"tree": []}
    return {"tree": []}

@app.post("/api/admin/organization-tree")
async def save_organization_tree(request: Request, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "organization.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    try:
        body = await request.json()
        tree = body["tree"] if isinstance(body, dict) and "tree" in body else body
        if not isinstance(tree, list):
            raise HTTPException(status_code=422, detail="tree must be an array")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Invalid JSON: {str(e)}")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM organization_tree")
    c.execute("INSERT INTO organization_tree (tree_data, updated_by, updated_at) VALUES (?, ?, ?)", (json.dumps(tree, ensure_ascii=False), username, dt.now().isoformat()))
    conn.commit()
    conn.close()
    return {"message": "Структура организации сохранена", "updated_by": username, "updated_at": dt.now().isoformat()}

# ============ УПРАВЛЕНИЕ AD ГРУППАМИ ============
@app.get("/api/admin/groups")
async def get_ad_groups(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "groups.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    groups = [dict(row) for row in c.execute("SELECT id, group_name, display_name, created_at FROM ad_groups ORDER BY display_name").fetchall()]
    conn.close()
    return {"groups": groups}

@app.post("/api/admin/groups")
async def add_ad_group(group_name: str = Form(...), display_name: str = Form(None), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "groups.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO ad_groups (group_name, display_name) VALUES (?, ?)", (group_name, display_name or group_name))
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Group already exists")
    finally:
        conn.close()
    return {"message": f"Group {group_name} added"}

@app.delete("/api/admin/groups/{group_id}")
async def delete_ad_group(group_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "groups.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM ad_groups WHERE id = ?", (group_id,))
    conn.commit()
    conn.close()
    return {"message": "Group deleted"}

# ============ УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ ============
@app.get("/api/admin/user-role")
async def get_current_user_role(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    role = await get_user_role(username)
    return {"role": role, "is_admin_by_group": is_admin_by_group(current_user.get("groups", []))}

@app.get("/api/admin/users")
async def get_users_with_roles(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "users.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_db()
    c = conn.cursor()
    db_users = c.execute("SELECT username, name FROM users ORDER BY name").fetchall()
    conn.close()
    
    conn2 = get_settings_db()
    c2 = conn2.cursor()
    
    user_roles = {}
    for row in c2.execute("SELECT ur.username, r.role_name FROM user_roles ur JOIN roles r ON ur.role_id = r.id").fetchall():
        user_roles[row[0]] = row[1]
    
    roles = [dict(row) for row in c2.execute("SELECT id, role_name, display_name FROM roles").fetchall()]
    conn2.close()
    
    users_list = []
    for db_user in db_users:
        users_list.append({
            "username": db_user[0],
            "name": db_user[1] or db_user[0],
            "email": "",
            "role": user_roles.get(db_user[0], "user"),
            "assigned_by": None
        })
    
    return {"users": users_list, "roles": roles}

@app.post("/api/admin/users/{username}/role")
async def assign_user_role(username: str, role_name: str = Form(...), current_user: dict = Depends(get_current_user)):
    admin_username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(admin_username, "users.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    role = c.execute("SELECT id FROM roles WHERE role_name = ?", (role_name,)).fetchone()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    c.execute("DELETE FROM user_roles WHERE username = ?", (username,))
    c.execute("INSERT INTO user_roles (username, role_id, assigned_by) VALUES (?, ?, ?)", (username, role[0], admin_username))
    conn.commit()
    conn.close()
    return {"message": f"Роль {role_name} назначена пользователю {username}"}

@app.get("/api/admin/ad-users")
async def search_ad_users(query: str = "", limit: int = 100, current_user: dict = Depends(get_current_user)):
    from server.ad_users import get_all_ad_users
    try:
        all_users = get_all_ad_users()
        try:
            query = urllib.parse.unquote(query)
        except:
            pass
        results = []
        if query and query.strip():
            query_lower = query.lower().strip()
            for user in all_users:
                ad_username = user.get('username', '').lower()
                display_name = user.get('name', '').lower()
                if query_lower in ad_username or query_lower in display_name:
                    results.append({'username': user.get('username'), 'display_name': user.get('name'), 'email': user.get('email')})
                    if len(results) >= limit:
                        break
        else:
            for user in all_users[:limit]:
                results.append({'username': user.get('username'), 'display_name': user.get('name'), 'email': user.get('email')})
        return {"users": results, "total": len(all_users)}
    except Exception as e:
        logger.error(f"Search error: {e}")
        return {"users": [], "total": 0, "error": str(e)}

@app.get("/api/users/authorized")
async def get_authorized_users(query: str = "", limit: int = 50, current_user: dict = Depends(get_current_user)):
    import urllib.parse
    conn = get_db()
    c = conn.cursor()
    
    try:
        query = urllib.parse.unquote(query)
    except:
        pass
    
    users = c.execute("""
        SELECT username, name as display_name, '' as email 
        FROM users 
        ORDER BY name
    """).fetchall()
    
    conn.close()
    
    result = []
    if query and query.strip():
        query_lower = query.lower().strip()
        for user in users:
            username = user[0].lower()
            display_name = (user[1] or user[0]).lower()
            if query_lower in username or query_lower in display_name:
                result.append({
                    "username": user[0],
                    "display_name": user[1] or user[0],
                    "email": user[2]
                })
                if len(result) >= limit:
                    break
    else:
        for user in users[:limit]:
            result.append({
                "username": user[0],
                "display_name": user[1] or user[0],
                "email": user[2]
            })
    
    return {"users": result}
    
# ============ УПРАВЛЕНИЕ КАТЕГОРИЯМИ РЕСУРСОВ ============
@app.get("/api/admin/resource-categories")
async def get_all_categories(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    categories = [dict(row) for row in c.execute("SELECT * FROM resource_categories ORDER BY sort_order, name").fetchall()]
    for cat in categories:
        cat['targets'] = [dict(row) for row in c.execute("SELECT id, target_type, target_id FROM category_targets WHERE category_id = ?", (cat['id'],)).fetchall()]
        cat['resources_count'] = c.execute("SELECT COUNT(*) FROM network_resources WHERE category_id = ?", (cat['id'],)).fetchone()[0]
    conn.close()
    return {"categories": categories}

@app.post("/api/admin/resource-categories")
async def add_category(name: str = Form(...), description: str = Form(None), icon: str = Form("📁"), is_global: int = Form(0), sort_order: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO resource_categories (name, description, icon, is_global, sort_order) VALUES (?, ?, ?, ?, ?)", (name, description, icon, is_global, sort_order))
    conn.commit()
    category_id = c.lastrowid
    conn.close()
    return {"id": category_id, "message": "Категория добавлена"}

@app.put("/api/admin/resource-categories/{category_id}")
async def update_category(category_id: int, name: str = Form(...), description: str = Form(None), icon: str = Form("📁"), is_global: int = Form(0), sort_order: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("UPDATE resource_categories SET name=?, description=?, icon=?, is_global=?, sort_order=? WHERE id=?", (name, description, icon, is_global, sort_order, category_id))
    conn.commit()
    conn.close()
    return {"message": "Категория обновлена"}

@app.delete("/api/admin/resource-categories/{category_id}")
async def delete_category(category_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM category_targets WHERE category_id = ?", (category_id,))
    c.execute("UPDATE network_resources SET category_id = NULL WHERE category_id = ?", (category_id,))
    c.execute("DELETE FROM resource_categories WHERE id = ?", (category_id,))
    conn.commit()
    conn.close()
    return {"message": "Категория удалена"}

@app.post("/api/admin/resource-categories/{category_id}/targets")
async def add_category_target(category_id: int, target_type: str = Form(...), target_id: str = Form(...), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    if not c.execute("SELECT id FROM resource_categories WHERE id = ?", (category_id,)).fetchone():
        raise HTTPException(status_code=404, detail="Category not found")
    c.execute("INSERT INTO category_targets (category_id, target_type, target_id) VALUES (?, ?, ?)", (category_id, target_type, target_id))
    conn.commit()
    conn.close()
    return {"message": "Target added successfully"}

@app.delete("/api/admin/resource-categories/{category_id}/targets/{target_id}")
async def delete_category_target(category_id: int, target_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "categories.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM category_targets WHERE id = ? AND category_id = ?", (target_id, category_id))
    conn.commit()
    conn.close()
    return {"message": "Target deleted successfully"}

# ============ УПРАВЛЕНИЕ СЕТЕВЫМИ РЕСУРСАМИ ============
def is_network_resource_visible_for_user(resource_id: int, username: str, user_groups: list, conn) -> bool:
    c = conn.cursor()
    resource = c.execute("SELECT is_global, category_id FROM network_resources WHERE id = ?", (resource_id,)).fetchone()
    if not resource:
        return False
    if resource[0] == 1:
        return True
    if c.execute("SELECT COUNT(*) FROM network_resource_targets WHERE resource_id = ?", (resource_id,)).fetchone()[0] > 0:
        for target_type, target_id in c.execute("SELECT target_type, target_id FROM network_resource_targets WHERE resource_id = ?", (resource_id,)).fetchall():
            if target_type == 'user' and target_id == username:
                return True
            elif target_type == 'group' and (c.execute("SELECT group_name FROM ad_groups WHERE id = ?", (target_id,)).fetchone() or [""])[0] in user_groups:
                return True
        return False
    if resource[1]:
        category = c.execute("SELECT is_global FROM resource_categories WHERE id = ?", (resource[1],)).fetchone()
        if category and category[0] == 1:
            return True
        for target_type, target_id in c.execute("SELECT target_type, target_id FROM category_targets WHERE category_id = ?", (resource[1],)).fetchall():
            if target_type == 'user' and target_id == username:
                return True
            elif target_type == 'group' and (c.execute("SELECT group_name FROM ad_groups WHERE id = ?", (target_id,)).fetchone() or [""])[0] in user_groups:
                return True
        return False
    creator = c.execute("SELECT created_by FROM network_resources WHERE id = ?", (resource_id,)).fetchone()
    return creator and creator[0] == username

@app.get("/api/network-resources")
async def get_user_network_resources(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    conn = get_settings_db()
    c = conn.cursor()
    all_resources = [dict(row) for row in c.execute("SELECT * FROM network_resources ORDER BY sort_order, resource_name").fetchall()]
    filtered_resources = []
    for r in all_resources:
        if is_network_resource_visible_for_user(r['id'], username, user_groups, conn):
            r['has_custom_targets'] = c.execute("SELECT COUNT(*) FROM network_resource_targets WHERE resource_id = ?", (r['id'],)).fetchone()[0] > 0
            r['inherits_from_category'] = not r['has_custom_targets'] and r.get('category_id') is not None
            filtered_resources.append(r)
    all_categories = [dict(row) for row in c.execute("SELECT * FROM resource_categories ORDER BY sort_order, name").fetchall()]
    result = []
    for cat in all_categories:
        if cat['is_global'] == 1:
            cat_visible = True
        else:
            cat_targets = c.execute("SELECT target_type, target_id FROM category_targets WHERE category_id = ?", (cat['id'],)).fetchall()
            cat_visible = False
            for target_type, target_id in cat_targets:
                if target_type == 'user' and target_id == username:
                    cat_visible = True
                    break
                elif target_type == 'group' and (c.execute("SELECT group_name FROM ad_groups WHERE id = ?", (target_id,)).fetchone() or [""])[0] in user_groups:
                    cat_visible = True
                    break
        if cat_visible:
            cat_resources = [r for r in filtered_resources if r.get('category_id') == cat['id']]
            cat_resources.sort(key=lambda x: x.get('sort_order', 0))
            result.append({'id': cat['id'], 'name': cat['name'], 'icon': cat.get('icon', '📁'), 'is_global': cat['is_global'], 'resources': cat_resources})
    uncategorized = [r for r in filtered_resources if not r.get('category_id')]
    if uncategorized:
        result.append({'id': None, 'name': 'Без категории', 'icon': '📁', 'is_global': 0, 'resources': uncategorized})
    conn.close()
    return {"categories": result}

@app.get("/api/admin/network-resources")
async def get_all_network_resources(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    resources = [dict(row) for row in c.execute("SELECT * FROM network_resources ORDER BY sort_order, resource_name").fetchall()]
    for res in resources:
        res['targets'] = [dict(row) for row in c.execute("SELECT id, target_type, target_id FROM network_resource_targets WHERE resource_id = ?", (res['id'],)).fetchall()]
        res['has_custom_targets'] = len(res['targets']) > 0
        if res.get('category_id') and not res['has_custom_targets']:
            res['inherits_from_category'] = True
            res['inherited_targets'] = [dict(row) for row in c.execute("SELECT id, target_type, target_id FROM category_targets WHERE category_id = ?", (res['category_id'],)).fetchall()]
        else:
            res['inherits_from_category'] = False
            res['inherited_targets'] = []
    conn.close()
    return {"resources": resources}

@app.post("/api/admin/network-resources")
async def add_network_resource(resource_name: str = Form(...), resource_path: str = Form(...), resource_type: str = Form("folder"), category_id: int = Form(None), is_global: int = Form(0), sort_order: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO network_resources (resource_name, resource_path, resource_type, category_id, is_global, sort_order, created_by) VALUES (?, ?, ?, ?, ?, ?, ?)",
              (resource_name, resource_path, resource_type, category_id if category_id and category_id > 0 else None, is_global, sort_order, username))
    conn.commit()
    resource_id = c.lastrowid
    conn.close()
    return {"id": resource_id, "message": "Ресурс добавлен"}

@app.put("/api/admin/network-resources/{resource_id}")
async def update_network_resource(resource_id: int, resource_name: str = Form(...), resource_path: str = Form(...), resource_type: str = Form("folder"), category_id: int = Form(None), is_global: int = Form(0), sort_order: int = Form(0), inherit_from_category: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("UPDATE network_resources SET resource_name=?, resource_path=?, resource_type=?, category_id=?, is_global=?, sort_order=? WHERE id=?",
              (resource_name, resource_path, resource_type, category_id if category_id and category_id > 0 else None, is_global, sort_order, resource_id))
    if inherit_from_category == 1:
        c.execute("DELETE FROM network_resource_targets WHERE resource_id = ?", (resource_id,))
    conn.commit()
    conn.close()
    return {"message": "Ресурс обновлен"}

@app.delete("/api/admin/network-resources/{resource_id}")
async def delete_network_resource(resource_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM network_resource_targets WHERE resource_id = ?", (resource_id,))
    c.execute("DELETE FROM network_resources WHERE id = ?", (resource_id,))
    conn.commit()
    conn.close()
    return {"message": "Ресурс удален"}

@app.post("/api/admin/network-resources/{resource_id}/targets")
async def add_network_resource_target(resource_id: int, target_type: str = Form(...), target_id: str = Form(...), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    if not c.execute("SELECT id FROM network_resources WHERE id = ?", (resource_id,)).fetchone():
        raise HTTPException(status_code=404, detail="Resource not found")
    c.execute("INSERT INTO network_resource_targets (resource_id, target_type, target_id) VALUES (?, ?, ?)", (resource_id, target_type, target_id))
    conn.commit()
    conn.close()
    return {"message": "Target added successfully"}

@app.delete("/api/admin/network-resources/{resource_id}/targets/{target_id}")
async def delete_network_resource_target(resource_id: int, target_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "network.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM network_resource_targets WHERE id = ? AND resource_id = ?", (target_id, resource_id))
    conn.commit()
    conn.close()
    return {"message": "Target deleted successfully"}

# ============ УПРАВЛЕНИЕ СЕРВИСАМИ ============
def is_service_visible_for_user(service_id: int, username: str, user_groups: list, conn) -> bool:
    c = conn.cursor()
    service = c.execute("SELECT is_global FROM services WHERE id = ?", (service_id,)).fetchone()
    if not service:
        return False
    if service[0] == 1:
        return True
    for target_type, target_id in c.execute("SELECT target_type, target_id FROM service_targets WHERE service_id = ?", (service_id,)).fetchall():
        if target_type == 'user' and target_id == username:
            return True
        elif target_type == 'group' and (c.execute("SELECT group_name FROM ad_groups WHERE id = ?", (target_id,)).fetchone() or [""])[0] in user_groups:
            return True
    return False

@app.get("/api/services")
async def get_user_services(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    conn = get_settings_db()
    c = conn.cursor()
    all_services = [dict(row) for row in c.execute("SELECT * FROM services").fetchall()]
    filtered_services = [s for s in all_services if is_service_visible_for_user(s['id'], username, user_groups, conn)]
    conn.close()
    return {"services": filtered_services}

@app.get("/api/admin/services")
async def get_all_services(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    services = [dict(row) for row in c.execute("SELECT * FROM services ORDER BY sort_order, service_name").fetchall()]
    for s in services:
        s['targets'] = [dict(row) for row in c.execute("SELECT id, target_type, target_id FROM service_targets WHERE service_id = ?", (s['id'],)).fetchall()]
    conn.close()
    return {"services": services}

@app.post("/api/admin/services")
async def add_service(service_name: str = Form(...), service_url: str = Form(...), service_icon: str = Form("🔗"), is_global: int = Form(0), sort_order: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO services (service_name, service_url, service_icon, is_global, sort_order) VALUES (?, ?, ?, ?, ?)", (service_name, service_url, service_icon, is_global, sort_order))
    conn.commit()
    service_id = c.lastrowid
    conn.close()
    return {"id": service_id, "message": "Сервис добавлен"}

@app.put("/api/admin/services/{service_id}")
async def update_service(service_id: int, service_name: str = Form(...), service_url: str = Form(...), service_icon: str = Form("🔗"), is_global: int = Form(0), sort_order: int = Form(0), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("UPDATE services SET service_name=?, service_url=?, service_icon=?, is_global=?, sort_order=? WHERE id=?", (service_name, service_url, service_icon, is_global, sort_order, service_id))
    conn.commit()
    conn.close()
    return {"message": "Сервис обновлен"}

@app.delete("/api/admin/services/{service_id}")
async def delete_service(service_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM service_targets WHERE service_id = ?", (service_id,))
    c.execute("DELETE FROM services WHERE id = ?", (service_id,))
    conn.commit()
    conn.close()
    return {"message": "Сервис удален"}

@app.post("/api/admin/services/{service_id}/targets")
async def add_service_target(service_id: int, target_type: str = Form(...), target_id: str = Form(...), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    if not c.execute("SELECT id FROM services WHERE id = ?", (service_id,)).fetchone():
        raise HTTPException(status_code=404, detail="Service not found")
    c.execute("INSERT INTO service_targets (service_id, target_type, target_id) VALUES (?, ?, ?)", (service_id, target_type, target_id))
    conn.commit()
    conn.close()
    return {"message": "Target added successfully"}

@app.delete("/api/admin/services/{service_id}/targets/{target_id}")
async def delete_service_target(service_id: int, target_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "services.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM service_targets WHERE id = ? AND service_id = ?", (target_id, service_id))
    conn.commit()
    conn.close()
    return {"message": "Target deleted successfully"}

# ============ API ДЛЯ IT-ЗАДАЧ ============
@app.get("/api/it-tasks")
async def get_user_it_tasks(archived: bool = False, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    if user_role == 'admin' or is_admin_by_group(user_groups):
        tasks = [dict(row) for row in c.execute("""
            SELECT t.*, tc.name as category_name, tc.color as category_color, et.name as equipment_name
            FROM it_tasks t
            LEFT JOIN task_categories tc ON t.category_id = tc.id
            LEFT JOIN equipment_types et ON t.equipment_id = et.id
            WHERE t.is_archived = ?
            ORDER BY t.created_date DESC
        """, (1 if archived else 0,)).fetchall()]
    else:
        tasks = [dict(row) for row in c.execute("""
            SELECT t.*, tc.name as category_name, tc.color as category_color, et.name as equipment_name
            FROM it_tasks t
            LEFT JOIN task_categories tc ON t.category_id = tc.id
            LEFT JOIN equipment_types et ON t.equipment_id = et.id
            WHERE t.is_archived = ? AND t.executor = ?
            ORDER BY t.created_date DESC
        """, (1 if archived else 0, username)).fetchall()]
    
    conn.close()
    return {"tasks": tasks}

@app.post("/api/it-tasks")
async def create_it_task(
    title: str = Form(...),
    description: str = Form(None),
    category_id: int = Form(None),
    equipment_id: int = Form(None),
    assigned_to: str = Form(None),
    components_status: str = Form("missing"),
    executor: str = Form(None),
    due_date: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    executor_email = None
    if executor:
        executor_email = get_user_email_by_name(executor)
    
    c.execute("""
        INSERT INTO it_tasks (title, description, category_id, equipment_id, assigned_to, 
                              components_status, executor, created_by, created_date, due_date, assigned_email)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (title, description, category_id, equipment_id, assigned_to, 
          components_status, executor, username, dt.now().isoformat(), due_date, executor_email))
    task_id = c.lastrowid
    conn.commit()
    conn.close()
    
    if executor and executor_email:
        send_task_email(executor_email, executor, title, description, due_date, task_id)
    elif executor and not executor_email:
        logger.warning(f"Не удалось найти email для исполнителя: {executor}")
    
    return {"id": task_id, "message": "Задача создана"}

@app.put("/api/it-tasks/{task_id}")
async def update_it_task(
    task_id: int,
    title: str = Form(None),
    description: str = Form(None),
    category_id: int = Form(None),
    equipment_id: int = Form(None),
    assigned_to: str = Form(None),
    components_status: str = Form(None),
    executor: str = Form(None),
    due_date: str = Form(None),
    is_archived: int = Form(None),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    updates = []
    values = []
    if title is not None:
        updates.append("title = ?"); values.append(title)
    if description is not None:
        updates.append("description = ?"); values.append(description)
    if category_id is not None:
        updates.append("category_id = ?"); values.append(category_id)
    if equipment_id is not None:
        updates.append("equipment_id = ?"); values.append(equipment_id)
    if assigned_to is not None:
        updates.append("assigned_to = ?"); values.append(assigned_to)
    if components_status is not None:
        updates.append("components_status = ?"); values.append(components_status)
    if executor is not None:
        updates.append("executor = ?"); values.append(executor)
        if executor:
            executor_email = get_user_email_by_name(executor)
            if executor_email:
                updates.append("assigned_email = ?"); values.append(executor_email)
    if due_date is not None:
        updates.append("due_date = ?"); values.append(due_date)
    if is_archived is not None:
        updates.append("is_archived = ?"); values.append(is_archived)
        if is_archived == 1:
            updates.append("completed_date = ?"); values.append(dt.now().isoformat())
    if updates:
        values.append(task_id)
        c.execute(f"UPDATE it_tasks SET {', '.join(updates)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return {"message": "Задача обновлена"}

@app.post("/api/it-tasks/{task_id}/restore")
async def restore_it_task(task_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("UPDATE it_tasks SET is_archived = 0, completed_date = NULL WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()
    return {"message": "Задача восстановлена из архива"}

@app.delete("/api/it-tasks/{task_id}")
async def delete_it_task(task_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM it_tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()
    return {"message": "Задача удалена"}

@app.post("/api/it-tasks/report")
async def generate_tasks_report(
    start_date: str = Form(...),
    end_date: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    tasks = c.execute("""
        SELECT t.*, tc.name as category_name
        FROM it_tasks t
        LEFT JOIN task_categories tc ON t.category_id = tc.id
        WHERE date(t.created_date) >= ? AND date(t.created_date) <= ?
        ORDER BY t.created_date DESC
    """, (start_date, end_date)).fetchall()
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по задачам"
    
    headers = ['ID', 'Название', 'Описание', 'Категория', 'Кому', 'Email', 'Исполнитель', 
               'Статус комплектующих', 'Дата создания', 'Срок', 'Дата выполнения', 'Архив']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_idx, task in enumerate(tasks, 2):
        task_dict = dict(task)
        ws.cell(row=row_idx, column=1, value=task_dict.get('id', ''))
        ws.cell(row=row_idx, column=2, value=task_dict.get('title', ''))
        ws.cell(row=row_idx, column=3, value=task_dict.get('description', '') or '')
        ws.cell(row=row_idx, column=4, value=task_dict.get('category_name', '') or '')
        ws.cell(row=row_idx, column=5, value=task_dict.get('assigned_to', '') or '')
        ws.cell(row=row_idx, column=6, value=task_dict.get('assigned_email', '') or '')
        ws.cell(row=row_idx, column=7, value=task_dict.get('executor', '') or '')
        
        status = task_dict.get('components_status', '')
        status_text = {'available': 'Есть', 'partial': 'Не полный', 'missing': 'Нет'}.get(status, status)
        ws.cell(row=row_idx, column=8, value=status_text)
        
        ws.cell(row=row_idx, column=9, value=task_dict.get('created_date', '') or '')
        ws.cell(row=row_idx, column=10, value=task_dict.get('due_date', '') or '')
        ws.cell(row=row_idx, column=11, value=task_dict.get('completed_date', '') or '')
        ws.cell(row=row_idx, column=12, value='Да' if task_dict.get('is_archived') else 'Нет')
    
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, len(tasks) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                try:
                    if len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tasks_report_{start_date}_{end_date}.xlsx"}
    )

@app.get("/api/task-categories")
async def get_task_categories_public(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    categories = [dict(row) for row in c.execute("SELECT * FROM task_categories ORDER BY sort_order, name").fetchall()]
    conn.close()
    return {"categories": categories}

# ============ API ДЛЯ КОМПЛЕКТУЮЩИХ ============
@app.get("/api/admin/equipment-types")
async def get_equipment_types(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "equipment.view", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("SELECT * FROM equipment_types ORDER BY name ASC")
    types = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"types": types}

@app.post("/api/admin/equipment-types")
async def add_equipment_type(name: str = Form(...), category: str = Form("pc_component"), unit: str = Form("шт"), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "equipment.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("INSERT INTO equipment_types (name, category, unit) VALUES (?, ?, ?)", (name, category, unit))
    conn.commit()
    type_id = c.lastrowid
    conn.close()
    return {"id": type_id, "message": "Тип добавлен"}

@app.put("/api/admin/equipment-types/{type_id}")
async def update_equipment_type(type_id: int, name: str = Form(...), category: str = Form(...), unit: str = Form(...), current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "equipment.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("UPDATE equipment_types SET name=?, category=?, unit=? WHERE id=?", (name, category, unit, type_id))
    conn.commit()
    conn.close()
    return {"message": "Тип обновлён"}

@app.delete("/api/admin/equipment-types/{type_id}")
async def delete_equipment_type(type_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not await has_permission(username, "equipment.manage", user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("DELETE FROM equipment_types WHERE id = ?", (type_id,))
    conn.commit()
    conn.close()
    return {"message": "Тип удалён"}

# ============ API ДЛЯ МОНИТОРИНГА IT-ЗАДАЧ ============
@app.get("/api/it-tasks/monitoring")
async def get_monitoring_tasks(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    monitoring = [dict(row) for row in c.execute("""
        SELECT m.*, tc.name as category_name, tc.color as category_color, et.name as equipment_name
        FROM it_task_monitoring m
        LEFT JOIN task_categories tc ON m.category_id = tc.id
        LEFT JOIN equipment_types et ON m.equipment_id = et.id
        ORDER BY m.created_at DESC
    """).fetchall()]
    
    conn.close()
    return {"monitoring_tasks": monitoring}

@app.post("/api/it-tasks/monitoring")
async def create_monitoring_task(request: Request, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    data = await request.json()
    
    interval_days = data.get("interval_days", 30)
    next_run = (dt.now() + timedelta(days=interval_days)).strftime("%Y-%m-%d")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("""
        INSERT INTO it_task_monitoring 
        (title_template, description_template, category_id, equipment_id, 
         assigned_to, executor, interval_days, is_active, next_run, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get("title_template"),
        data.get("description_template"),
        data.get("category_id"),
        data.get("equipment_id"),
        data.get("assigned_to"),
        data.get("executor"),
        interval_days,
        1 if data.get("is_active") else 0,
        next_run,
        username
    ))
    
    conn.commit()
    monitoring_id = c.lastrowid
    conn.close()
    
    return {"id": monitoring_id, "message": "Мониторинг создан"}

@app.put("/api/it-tasks/monitoring/{monitoring_id}")
async def update_monitoring_task(monitoring_id: int, request: Request, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    data = await request.json()
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("""
        UPDATE it_task_monitoring 
        SET title_template = ?, description_template = ?, category_id = ?, 
            equipment_id = ?, assigned_to = ?, executor = ?, 
            interval_days = ?, is_active = ?
        WHERE id = ?
    """, (
        data.get("title_template"),
        data.get("description_template"),
        data.get("category_id"),
        data.get("equipment_id"),
        data.get("assigned_to"),
        data.get("executor"),
        data.get("interval_days", 30),
        1 if data.get("is_active") else 0,
        monitoring_id
    ))
    
    conn.commit()
    conn.close()
    
    return {"message": "Мониторинг обновлён"}

@app.delete("/api/it-tasks/monitoring/{monitoring_id}")
async def delete_monitoring_task(monitoring_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("SELECT id FROM it_task_monitoring WHERE id = ?", (monitoring_id,))
    if not c.fetchone():
        raise HTTPException(status_code=404, detail="Мониторинг не найден")
    
    c.execute("DELETE FROM it_task_monitoring WHERE id = ?", (monitoring_id,))
    conn.commit()
    conn.close()
    
    return {"message": "Мониторинг удалён"}

@app.post("/api/it-tasks/monitoring/{monitoring_id}/toggle")
async def toggle_monitoring(monitoring_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("SELECT is_active FROM it_task_monitoring WHERE id = ?", (monitoring_id,))
    row = c.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Мониторинг не найден")
    
    new_status = 0 if row[0] else 1
    c.execute("UPDATE it_task_monitoring SET is_active = ? WHERE id = ?", (new_status, monitoring_id))
    conn.commit()
    conn.close()
    
    return {"is_active": new_status, "message": "Статус изменён"}

@app.post("/api/it-tasks/monitoring/{monitoring_id}/run")
async def run_monitoring_now(monitoring_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    user_role = await get_user_role(username)
    
    if user_role not in ['admin', 'it_engineer'] and not is_admin_by_group(user_groups):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute("""
        SELECT title_template, description_template, category_id, equipment_id,
               assigned_to, executor, interval_days
        FROM it_task_monitoring WHERE id = ?
    """, (monitoring_id,))
    row = c.fetchone()
    
    if not row:
        raise HTTPException(status_code=404, detail="Мониторинг не найден")
    
    title = row[0]
    if "{date}" in title:
        title = title.replace("{date}", dt.now().strftime("%d.%m.%Y"))
    if "{DATE}" in title:
        title = title.replace("{DATE}", dt.now().strftime("%d.%m.%Y"))
    
    executor_email = None
    if row[5]:
        executor_email = get_user_email_by_name(row[5])
    
    c.execute("""
        INSERT INTO it_tasks 
        (title, description, category_id, equipment_id, assigned_to, executor, 
         components_status, monitoring_id, created_by, created_date, assigned_email)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        title,
        row[1],
        row[2],
        row[3],
        row[4],
        row[5],
        "missing",
        monitoring_id,
        username,
        dt.now().isoformat(),
        executor_email
    ))
    
    task_id = c.lastrowid
    
    next_run = (dt.now() + timedelta(days=row[6])).strftime("%Y-%m-%d")
    c.execute("""
        UPDATE it_task_monitoring 
        SET last_run = date('now'), next_run = ?
        WHERE id = ?
    """, (next_run, monitoring_id))
    
    conn.commit()
    conn.close()
    
    if row[5] and executor_email:
        send_task_email(executor_email, row[5], title, row[1], None, task_id)
    
    return {"task_id": task_id, "message": "Задача создана"}

# ============ АДМИНИСТРИРОВАНИЕ ============
@app.get("/api/admin/reset-chats")
async def reset_all_chats(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    user_groups = current_user.get("groups", [])
    if not is_admin_by_group(user_groups) and await get_user_role(username) != 'admin':
        raise HTTPException(status_code=403, detail="Forbidden")
    with get_db() as conn:
        conn.execute("DELETE FROM messages")
        conn.execute("DELETE FROM chat_members")
        conn.execute("DELETE FROM chats")
        conn.execute("DELETE FROM read_receipts")
        conn.commit()
        cursor = conn.execute("INSERT INTO chats (name, created_by, is_general, created_at) VALUES (?, ?, ?, ?)", ("Общий чат", "system", 1, dt.now().isoformat()))
        conn.commit()
        for user in conn.execute("SELECT username FROM users").fetchall():
            conn.execute("INSERT OR IGNORE INTO chat_members (chat_id, username) VALUES (?, ?)", (cursor.lastrowid, user["username"]))
        conn.commit()
    return {"success": True, "message": "Все чаты сброшены"}

@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": dt.now().isoformat()}

# ============ API ДЛЯ ТОСТ-УВЕДОМЛЕНИЙ ============

@app.get("/api/notifications/toast")
async def get_toast_notifications(current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    
    conn = get_settings_db()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS toast_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT NOT NULL,
        event_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        event_date TEXT,
        event_time TEXT,
        location TEXT,
        description TEXT,
        event_type TEXT,
        remind_before INTEGER,
        shown INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (event_id) REFERENCES calendar_events(id) ON DELETE CASCADE
    )''')
    conn.commit()
    
    c.execute("""
        SELECT id, event_id, title, event_date, event_time, location, description, event_type, remind_before
        FROM toast_notifications
        WHERE user_id = ? AND shown = 0
        ORDER BY created_at ASC
    """, (username,))
    
    notifications = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return {"notifications": notifications}

@app.put("/api/notifications/toast/{notification_id}/shown")
async def mark_toast_shown(notification_id: int, current_user: dict = Depends(get_current_user)):
    username = current_user.get("sub")
    
    conn = get_settings_db()
    c = conn.cursor()
    c.execute("""
        UPDATE toast_notifications
        SET shown = 1
        WHERE id = ? AND user_id = ?
    """, (notification_id, username))
    conn.commit()
    conn.close()
    
    return {"message": "ok"}

@app.on_event("startup")
async def start_reminder_task():
    asyncio.create_task(check_and_send_reminders())
    logger.info("✅ Фоновая задача отправки напоминаний запущена")

# ============ ЗАПУСК ============
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)